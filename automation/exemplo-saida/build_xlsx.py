import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Situacao por Estado"

headers = [
    "Estado",
    "Governador atual / Grupo situação",
    "Partido",
    "Elegível reeleição",
    "Nome a buscar nas pesquisas",
    "Confiança do nome",
    "Observação",
]

data = [
["AC","Mailza Assis (assumiu após saída de Gladson Cameli)","PP","Não se aplica (sucessão)","Mailza Assis","Provável","Cameli renunciou ao Senado; Mailza assumiu e é pré-candidata à reeleição com seu apoio."],
["AL","Paulo Dantas","MDB","Não vai concorrer","Renan Filho","Provável","Dantas descarta candidatura e apoia Renan Filho (MDB), ex-ministro dos Transportes."],
["AP","Clécio Luís","União Brasil","Sim","Clécio Luís","Provável","Governador confirmou pré-candidatura à reeleição."],
["AM","Roberto Cidade (assumiu após saída de Wilson Lima)","União Brasil","Não se aplica (sucessão)","Roberto Cidade","Provável","Lima renunciou ao Senado; Cidade foi eleito indiretamente e é o sucessor confirmado."],
["BA","Jerônimo Rodrigues","PT","Sim","Jerônimo Rodrigues","Provável","Busca reeleição; ACM Neto (União Brasil) desponta como principal adversário."],
["CE","Elmano de Freitas","PT","Sim","Elmano de Freitas","Provável","Confirmado à reeleição; Ciro Gomes (PSDB) é o principal adversário."],
["DF","Celina Leão (assumiu após saída de Ibaneis Rocha)","PP","Não se aplica (sucessão)","Celina Leão","Provável","Ibaneis renunciou ao Senado; Celina assumiu e concorre com apoio bolsonarista."],
["ES","Ricardo Ferraço (assumiu após saída de Renato Casagrande)","MDB","Não se aplica (sucessão)","Ricardo Ferraço","Provável","Casagrande renunciou ao Senado; Ferraço é o candidato de continuidade do grupo."],
["GO","Daniel Vilela (assumiu após saída de Ronaldo Caiado)","MDB","Não se aplica (sucessão)","Daniel Vilela","Provável","Caiado renunciou para concorrer à Presidência pelo PSD; Vilela busca reeleição."],
["MA","Carlos Brandão","sem partido","Não vai concorrer","Orleans Brandão","Provável","Brandão cumpre o mandato até o fim; sobrinho Orleans Brandão (MDB) é candidato do grupo."],
["MT","Otaviano Pivetta (assumiu após saída de Mauro Mendes)","Republicanos","Não se aplica (sucessão)","Otaviano Pivetta","Provável","Mendes renunciou ao Senado; Pivetta é pré-candidato à reeleição."],
["MS","Eduardo Riedel","PP","Sim","Eduardo Riedel","Provável","Confirmado à reeleição."],
["MG","Mateus Simões (assumiu após saída de Romeu Zema)","PSD","Não se aplica (sucessão)","Mateus Simões","Provável","Zema renunciou para disputar a Presidência pelo Novo; Simões é candidato de continuidade."],
["PA","Hana Ghassan (assumiu após saída de Helder Barbalho)","MDB","Não se aplica (sucessão)","Hana Ghassan","Provável","Barbalho renunciou ao Senado; Hana lidera as pesquisas para o governo estadual."],
["PB","Lucas Ribeiro (assumiu após saída de João Azevêdo)","PP","Não se aplica (sucessão)","Lucas Ribeiro","Provável","Azevêdo renunciou ao Senado; Ribeiro assumiu e concorre à reeleição."],
["PR","Ratinho Junior","PSD","Não vai concorrer","Sandro Alex","Especulativo","Ratinho Jr. cumpre o mandato; indicou Sandro Alex, ainda pouco conhecido, como sucessor."],
["PE","Raquel Lyra","PSD","Sim","Raquel Lyra","Provável","Confirmada à reeleição."],
["PI","Rafael Fonteles","PT","Sim","Rafael Fonteles","Provável","Busca reeleição; chapa majoritária ainda em definição."],
["RJ","Douglas Ruas (assumiu após crise sucessória e renúncia de Cláudio Castro)","PL","Não se aplica (sucessão)","Douglas Ruas","Especulativo","Castro renunciou e ficou inelegível; Ruas assumiu via Alerj e foi oficializado candidato pelo PL."],
["RN","Fátima Bezerra","PT","Não vai concorrer","Cadu Xavier","Provável","Fátima decidiu cumprir o mandato até o fim; indicou Cadu Xavier (PT) como sucessor."],
["RS","Eduardo Leite","PSD","Não vai concorrer","Gabriel Souza","Provável","Leite não disputa cargo em 2026; apoia o vice Gabriel Souza (MDB) na sucessão."],
["RO","Marcos Rocha","PSD","Não vai concorrer","Adailton Fúria","Especulativo","Rocha descartou renúncia; indicou o prefeito de Cacoal, Fúria, como seu sucessor."],
["RR","Arthur Henrique (eleito na eleição suplementar, resultado sub judice)","PL","Eleição suplementar","Arthur Henrique","Imprevisível","Denarium ficou inelegível e Damião foi cassado pelo TSE; resultado de junho segue sub judice."],
["SC","Jorginho Mello","PL","Sim","Jorginho Mello","Provável","Confirmou pré-candidatura à reeleição."],
["SP","Tarcísio de Freitas","Republicanos","Sim","Tarcísio de Freitas","Provável","Desistiu da Presidência após entrada de Flávio Bolsonaro; concorre à reeleição."],
["SE","Fábio Mitidieri","PSD","Sim","Fábio Mitidieri","Provável","Confirmou pré-candidatura à reeleição."],
["TO","Wanderlei Barbosa","Republicanos","Não vai concorrer","Professora Dorinha","Provável","Barbosa não terá candidato próprio; apoia a senadora Dorinha (União Brasil) ao governo."],
]

assert len(data) == 27, len(data)

font_name = "Arial"
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name=font_name, bold=True, color="FFFFFF", size=11)
body_font = Font(name=font_name, size=10)
wrap = Alignment(wrap_text=True, vertical="top")

ws.append(headers)
for col in range(1, len(headers)+1):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

for row in data:
    ws.append(row)

for r in range(2, 2+len(data)):
    for col in range(1, len(headers)+1):
        c = ws.cell(row=r, column=col)
        c.font = body_font
        c.alignment = wrap

col_widths = [8, 42, 14, 22, 24, 16, 55]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{1+len(data)}"

wb.save("aba1_situacao_por_estado.xlsx")
print("saved", len(data), "rows")
